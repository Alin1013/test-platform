from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_HEADERS = [
    "用例目录",
    "用例名称",
    "需求ID",
    "前置条件",
    "用例步骤",
    "预期结果",
    "用例类型",
    "用例状态",
    "用例等级",
    "创建人",
    "归属迭代",
]


def _join(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(item) for item in value if str(item).strip())
    if isinstance(value, dict):
        return "\n".join(f"{key}: {val}" for key, val in value.items() if str(val).strip())
    return str(value)


def _render_steps(steps: Any) -> str:
    if not isinstance(steps, list):
        return _join(steps)

    rendered = []
    for index, step in enumerate(steps, 1):
        if isinstance(step, dict):
            action = step.get("action", "")
            expected = step.get("expected", "")
            rendered.append(f"{step.get('index', index)}. {action}")
            if expected:
                rendered.append(f"   预期：{expected}")
        else:
            rendered.append(f"{index}. {step}")
    return "\n".join(rendered)


def _cell_value(value: Any):
    text = _join(value).strip()
    return text or None


class TemplateService:
    @staticmethod
    def default_schema(name: str = "默认模板") -> Dict[str, Any]:
        return {
            "name": name,
            "headers": DEFAULT_HEADERS,
            "example_rows": [],
            "static_columns": {},
        }

    @staticmethod
    def parse_template_bytes(data: bytes, filename: str = "template.xlsx") -> Dict[str, Any]:
        workbook = load_workbook(BytesIO(data), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header_row = next((row for row in rows if any(value not in (None, "") for value in row)), None)
        if header_row is None:
            return TemplateService.default_schema(Path(filename).stem or "默认模板")

        headers = [str(value).strip() for value in header_row if value not in (None, "")]
        header_index = rows.index(header_row)
        example_rows = []
        for row in rows[header_index + 1: header_index + 4]:
            if not any(value not in (None, "") for value in row):
                continue
            example_rows.append({
                headers[index]: row[index] if row[index] is not None else ""
                for index in range(min(len(headers), len(row)))
            })

        return {
            "name": Path(filename).stem or "默认模板",
            "headers": headers or DEFAULT_HEADERS,
            "example_rows": example_rows,
            "static_columns": {},
        }

    @staticmethod
    def build_row(case: Dict[str, Any], defaults: Dict[str, Any], headers: List[str] | None = None) -> Dict[str, str]:
        mapped = {
            "用例目录": _join(case.get("catalog")),
            "用例名称": _join(case.get("title") or case.get("name")),
            "需求ID": _join(defaults.get("requirement_ids") or case.get("requirement_ids")),
            "前置条件": _join(case.get("preconditions") or case.get("precondition")),
            "用例步骤": _render_steps(case.get("steps") or case.get("test_steps")),
            "预期结果": _join(case.get("expected_result") or case.get("expected")),
            "用例类型": _join(defaults.get("case_type") or case.get("case_type")),
            "用例状态": "",
            "用例等级": _join(case.get("priority") or case.get("level")),
            "创建人": _join(defaults.get("case_creator") or case.get("creator")),
            "归属迭代": _join(defaults.get("iteration") or case.get("iteration")),
        }

        for header in headers or []:
            mapped.setdefault(header, _join(case.get(header)))
        return mapped

    @staticmethod
    def build_rows(cases: List[Dict[str, Any]], defaults: Dict[str, Any], headers: List[str] | None = None) -> List[Dict[str, str]]:
        return [TemplateService.build_row(case, defaults, headers) for case in cases]

    @staticmethod
    def build_workbook_bytes(
        schema: Dict[str, Any] | None,
        cases: List[Dict[str, Any]],
        defaults: Dict[str, Any],
    ) -> bytes:
        schema = schema or TemplateService.default_schema()
        headers = schema.get("headers") or DEFAULT_HEADERS
        static_columns = schema.get("static_columns") or {}

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "测试用例"
        sheet.append(headers)

        for case in cases:
            row = TemplateService.build_row(case, defaults, headers)
            values = []
            for header in headers:
                if header == "用例状态":
                    values.append(None)
                elif header in static_columns:
                    values.append(_cell_value(static_columns[header]))
                else:
                    values.append(_cell_value(row.get(header)))
            sheet.append(values)

        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_cells in sheet.columns:
            header = str(column_cells[0].value or "")
            width = min(max(len(header) + 6, 14), 36)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
            for cell in column_cells[1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
