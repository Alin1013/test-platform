from asgiref.sync import async_to_sync
import json
import zipfile
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from rest_framework.test import APIRequestFactory
from io import BytesIO
from unittest import mock
from unittest.mock import AsyncMock
from PIL import Image
from openpyxl import Workbook, load_workbook

from apps.requirement_analysis.document_parser import DocumentParser, UnsupportedSourceFileError
from apps.requirement_analysis.generation_pipeline import (
    PRD2CasePipeline,
    VisionDocumentExtractor,
    build_excel_rows,
    parse_json_payload,
)
from apps.requirement_analysis.models import (
    AIModelConfig,
    PromptConfig,
    TestCaseGenerationTask,
    TestCaseTemplateConfig,
)
from apps.requirement_analysis.serializers import TestCaseGenerationRequestSerializer
from apps.requirement_analysis.template_service import TemplateService
from apps.requirement_analysis.views import TestCaseGenerationTaskViewSet
from apps.projects.models import Project
from apps.testcases.models import TestCase as ManagedTestCase


def make_xmind_bytes():
    content = [
        {
            "title": "Sheet 1",
            "rootTopic": {
                "title": "登录模块测试点",
                "children": {
                    "attached": [
                        {
                            "title": "验证码登录",
                            "children": {
                                "attached": [
                                    {"title": "正确手机号和验证码可以登录成功"},
                                    {"title": "验证码错误时提示登录失败"},
                                ]
                            },
                        }
                    ]
                },
            },
        }
    ]
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("content.json", json.dumps(content, ensure_ascii=False))
    return buffer.getvalue()


class PRD2CaseTaskModelTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="prd2case_user",
            password="password123",
        )

    def test_task_defaults_include_review_artifacts(self):
        task = TestCaseGenerationTask.objects.create(
            task_id="TASK_TEST001",
            title="登录 PRD",
            requirement_text="手机号验证码登录需求",
            requirement_ids=["REQ-LOGIN-001"],
            case_type="功能测试",
            case_creator="张三",
            iteration="2026.06",
            created_by=self.user,
        )

        self.assertEqual(task.requirement_ids, ["REQ-LOGIN-001"])
        self.assertEqual(task.case_type, "功能测试")
        self.assertEqual(task.case_creator, "张三")
        self.assertEqual(task.iteration, "2026.06")
        self.assertEqual(task.test_points, [])
        self.assertEqual(task.test_cases_json, [])
        self.assertEqual(task.test_points_review_status, "pending")
        self.assertEqual(task.test_cases_review_status, "pending")

    def test_ai_model_choices_include_openai_vision(self):
        model_type_field = AIModelConfig._meta.get_field("model_type")
        role_field = AIModelConfig._meta.get_field("role")

        self.assertIn(("openai", "OpenAI"), model_type_field.choices)
        self.assertIn(("vision", "视觉解析模型"), role_field.choices)

    def test_task_defaults_include_source_and_template_fields(self):
        task = TestCaseGenerationTask.objects.create(
            task_id="TASK_REF001",
            title="上传 PRD",
            requirement_text="",
            requirement_ids=["REQ-1"],
            case_type="功能测试",
            case_creator="张三",
            iteration="2026.06",
            created_by=self.user,
        )

        self.assertEqual(task.source_file_type, "")
        self.assertEqual(task.source_extract_status, "pending")
        self.assertEqual(task.source_extract_error, "")
        self.assertEqual(task.template_schema, {})
        self.assertEqual(task.selected_template_name, "")

    def test_template_config_model_defaults(self):
        config = TestCaseTemplateConfig.objects.create(
            name="默认模板",
            template_schema={"headers": ["用例名称"]},
            created_by=self.user,
        )

        self.assertTrue(config.is_active)
        self.assertEqual(config.template_schema["headers"], ["用例名称"])


class TestCaseGenerationRequestSerializerTests(TestCase):
    def test_required_prd_metadata_is_validated(self):
        serializer = TestCaseGenerationRequestSerializer(data={
            "title": "登录 PRD",
            "requirement_text": "手机号验证码登录需求",
            "use_writer_model": True,
            "use_reviewer_model": True,
        })

        self.assertFalse(serializer.is_valid())
        self.assertIn("requirement_ids", serializer.errors)
        self.assertIn("case_type", serializer.errors)
        self.assertIn("case_creator", serializer.errors)
        self.assertIn("iteration", serializer.errors)

    def test_requirement_ids_string_is_normalized_to_list(self):
        serializer = TestCaseGenerationRequestSerializer(data={
            "title": "登录 PRD",
            "requirement_text": "手机号验证码登录需求",
            "requirement_ids": "REQ-1, REQ-2",
            "case_type": "功能测试",
            "case_creator": "张三",
            "iteration": "2026.06",
            "use_writer_model": True,
            "use_reviewer_model": False,
        })

        self.assertTrue(serializer.is_valid(), serializer.errors)
        self.assertEqual(serializer.validated_data["requirement_ids"], ["REQ-1", "REQ-2"])


class DocumentParserTests(TestCase):
    def test_html_parser_preserves_text_and_tables(self):
        html = b"<h1>Login Requirement</h1><table><tr><th>Field</th><td>Phone</td></tr></table>"

        result = DocumentParser.extract_from_bytes("prd.html", html)

        self.assertIn("Login Requirement", result.text)
        self.assertIn("Field", result.text)
        self.assertIn("Phone", result.text)
        self.assertEqual(result.file_type, "html")

    def test_markdown_parser_treats_content_as_text(self):
        markdown = "# Login Requirement\n\n- Phone login\n- SMS code"

        result = DocumentParser.extract_from_bytes("prd.md", markdown.encode("utf-8"))

        self.assertIn("Login Requirement", result.text)
        self.assertIn("Phone login", result.text)
        self.assertEqual(result.file_type, "md")

    def test_unsupported_file_type_raises_clear_error(self):
        with self.assertRaises(UnsupportedSourceFileError):
            DocumentParser.extract_from_bytes("prd.zip", b"bad")

    def test_bmp_image_uses_vision_extractor(self):
        image = Image.new("RGB", (4, 4), "white")
        buffer = BytesIO()
        image.save(buffer, format="BMP")
        calls = []

        def fake_vision(filename, content_type, data):
            calls.append((filename, content_type, data[:8]))
            return "图片中的登录 PRD"

        result = DocumentParser.extract_from_bytes(
            "prd.bmp",
            buffer.getvalue(),
            vision_extractor=fake_vision,
        )

        self.assertEqual(result.text, "图片中的登录 PRD")
        self.assertEqual(result.file_type, "bmp")
        self.assertEqual(calls[0][1], "image/png")

    def test_xmind_parser_extracts_leaf_topics_as_test_points(self):
        result = DocumentParser.extract_from_bytes("login.xmind", make_xmind_bytes())

        self.assertEqual(result.file_type, "xmind")
        self.assertIn("正确手机号和验证码可以登录成功", result.text)
        points = result.metadata["xmind_test_points"]
        self.assertEqual(len(points), 2)
        self.assertEqual(points[0]["title"], "正确手机号和验证码可以登录成功")
        self.assertEqual(points[0]["source_trace"], "登录模块测试点 > 验证码登录 > 正确手机号和验证码可以登录成功")


class TemplateServiceTests(TestCase):
    def test_template_schema_reads_headers_and_example_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["用例目录", "用例名称", "用例状态"])
        sheet.append(["登录", "登录成功", ""])
        buffer = BytesIO()
        workbook.save(buffer)

        schema = TemplateService.parse_template_bytes(buffer.getvalue(), filename="template.xlsx")

        self.assertEqual(schema["headers"], ["用例目录", "用例名称", "用例状态"])
        self.assertEqual(schema["example_rows"][0]["用例名称"], "登录成功")

    def test_export_xlsx_keeps_case_status_blank(self):
        data = TemplateService.build_workbook_bytes(
            schema={"headers": ["用例名称", "用例状态", "创建人"]},
            cases=[{"title": "登录成功", "case_status": "已评审", "creator": "李四"}],
            defaults={"case_creator": "张三"},
        )

        workbook = load_workbook(BytesIO(data))
        row = [cell.value for cell in workbook.active[2]]

        self.assertEqual(row, ["登录成功", None, "张三"])


class PRD2CasePipelineRevisionTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("pipeline_user", password="password123")
        self.writer_config = AIModelConfig.objects.create(
            name="Writer",
            model_type="deepseek",
            role="writer",
            api_key="test-key",
            base_url="https://api.example.com/v1",
            model_name="writer-model",
            created_by=self.user,
        )
        self.prompt = PromptConfig.objects.create(
            name="Writer prompt",
            prompt_type="writer",
            content="你是测试专家",
            is_active=True,
            created_by=self.user,
        )
        self.task = TestCaseGenerationTask.objects.create(
            task_id="TASK_PIPE001",
            title="登录 PRD",
            requirement_text="手机号验证码登录",
            requirement_ids=["REQ-1"],
            case_type="功能测试",
            case_creator="张三",
            iteration="2026.06",
            test_points=[{"id": "TP-1", "title": "旧点"}],
            test_cases_json=[{"id": "TC-1", "title": "旧用例"}],
            template_schema={"headers": ["用例名称", "用例状态", "创建人"]},
            writer_model_config=self.writer_config,
            writer_prompt_config=self.prompt,
            created_by=self.user,
        )

    def test_revise_test_points_returns_new_version(self):
        with mock.patch(
            "apps.requirement_analysis.generation_pipeline.AIModelService.call_openai_compatible_api",
            new=AsyncMock(return_value={
                "choices": [{"message": {"content": '[{"id":"TP-1","title":"新点"}]'}}],
            }),
        ):
            result = async_to_sync(PRD2CasePipeline(self.task).revise_test_points)("补充异常场景")

        self.assertEqual(result.artifact[0]["title"], "新点")

    def test_generate_test_points_does_not_lazy_load_config_in_async_context(self):
        task = TestCaseGenerationTask.objects.get(task_id=self.task.task_id)
        mocked_call = AsyncMock(return_value={
            "choices": [{"message": {"content": '[{"id":"TP-1","title":"新点"}]'}}],
        })

        with mock.patch(
            "apps.requirement_analysis.generation_pipeline.AIModelService.call_openai_compatible_api",
            new=mocked_call,
        ):
            result = async_to_sync(PRD2CasePipeline(task).generate_test_points)()

        self.assertEqual(result.artifact[0]["title"], "新点")

    def test_revise_test_cases_includes_template_schema(self):
        mocked_call = AsyncMock(return_value={
            "choices": [{"message": {"content": '[{"id":"TC-1","title":"新用例"}]'}}],
        })
        with mock.patch(
            "apps.requirement_analysis.generation_pipeline.AIModelService.call_openai_compatible_api",
            new=mocked_call,
        ):
            result = async_to_sync(PRD2CasePipeline(self.task).revise_test_cases)("补充失败分支")

        messages = mocked_call.call_args.args[1]
        self.assertIn("模板结构", messages[1]["content"])
        self.assertEqual(result.artifact[0]["title"], "新用例")

    def test_generate_cases_from_points_does_not_lazy_load_config_in_async_context(self):
        task = TestCaseGenerationTask.objects.get(task_id=self.task.task_id)
        mocked_call = AsyncMock(return_value={
            "choices": [{"message": {"content": '[{"id":"TC-1","title":"新用例"}]'}}],
        })

        with mock.patch(
            "apps.requirement_analysis.generation_pipeline.AIModelService.call_openai_compatible_api",
            new=mocked_call,
        ):
            result = async_to_sync(PRD2CasePipeline(task).generate_cases_from_points)()

        self.assertEqual(result.artifact[0]["title"], "新用例")

    def test_generate_cases_from_points_batches_large_point_sets(self):
        task = TestCaseGenerationTask.objects.get(task_id=self.task.task_id)
        task.test_points = [
            {"id": f"TP-{index:03d}", "title": f"测试点 {index:03d}"}
            for index in range(1, 26)
        ]
        task.save(update_fields=["test_points", "updated_at"])
        mocked_call = AsyncMock(side_effect=[
            {"choices": [{"message": {"content": '[{"id":"TC-1","title":"第一批用例"}]'}}]},
            {"choices": [{"message": {"content": '[{"id":"TC-2","title":"第二批用例"}]'}}]},
        ])

        with mock.patch(
            "apps.requirement_analysis.generation_pipeline.AIModelService.call_openai_compatible_api",
            new=mocked_call,
        ):
            result = async_to_sync(PRD2CasePipeline(task).generate_cases_from_points)()

        self.assertEqual(mocked_call.call_count, 2)
        first_prompt = mocked_call.call_args_list[0].args[1][1]["content"]
        second_prompt = mocked_call.call_args_list[1].args[1][1]["content"]
        self.assertIn("TP-001", first_prompt)
        self.assertNotIn("TP-021", first_prompt)
        self.assertIn("TP-021", second_prompt)
        self.assertEqual([case["title"] for case in result.artifact], ["第一批用例", "第二批用例"])

    def test_vision_document_extractor_sends_image_payload(self):
        vision_config = AIModelConfig.objects.create(
            name="Vision",
            model_type="openai",
            role="vision",
            api_key="test-key",
            base_url="https://api.openai.com/v1",
            model_name="gpt-4.1-mini",
            created_by=self.user,
        )
        mocked_call = AsyncMock(return_value={
            "choices": [{"message": {"content": "图片 PRD 文本"}}],
        })

        with mock.patch(
            "apps.requirement_analysis.generation_pipeline.AIModelService.call_openai_compatible_api",
            new=mocked_call,
        ):
            text = async_to_sync(VisionDocumentExtractor.extract_text)(
                vision_config,
                "prd.png",
                "image/png",
                b"image-bytes",
            )

        messages = mocked_call.call_args.args[1]
        self.assertEqual(text, "图片 PRD 文本")
        self.assertIsInstance(messages[1]["content"], list)
        self.assertIn("data:image/png;base64,", messages[1]["content"][1]["image_url"]["url"])


class PRD2CasePipelineUtilityTests(TestCase):
    def test_parse_json_payload_extracts_fenced_json(self):
        payload = '```json\n[{"id":"TP-001","title":"登录主流程"}]\n```'

        self.assertEqual(parse_json_payload(payload), [{"id": "TP-001", "title": "登录主流程"}])

    def test_build_excel_rows_keeps_case_status_blank(self):
        rows = build_excel_rows(
            [{
                "catalog": "登录",
                "title": "验证码登录成功",
                "requirement_ids": ["REQ-1"],
                "preconditions": ["用户未登录"],
                "steps": [{"index": 1, "action": "输入验证码", "expected": "登录成功"}],
                "expected_result": "登录成功",
                "case_type": "功能测试",
                "priority": "P1",
                "creator": "张三",
                "iteration": "2026.06",
            }],
            defaults={
                "requirement_ids": ["REQ-DEFAULT"],
                "case_type": "功能测试",
                "case_creator": "张三",
                "iteration": "2026.06",
            },
        )

        self.assertEqual(rows[0]["用例状态"], "")
        self.assertEqual(rows[0]["需求ID"], "REQ-DEFAULT")
        self.assertEqual(rows[0]["创建人"], "张三")
        self.assertEqual(rows[0]["归属迭代"], "2026.06")
        self.assertIn("1. 输入验证码", rows[0]["用例步骤"])


class TestCaseGenerationConfigValidationTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("api_user", password="password123")
        self.factory = APIRequestFactory()

    def test_generate_requires_writer_api_key(self):
        AIModelConfig.objects.create(
            name="Writer without key",
            model_type="deepseek",
            role="writer",
            api_key="",
            base_url="https://api.example.com",
            model_name="model",
            is_active=True,
            created_by=self.user,
        )
        PromptConfig.objects.create(
            name="Writer prompt",
            prompt_type="writer",
            content="你是测试专家",
            is_active=True,
            created_by=self.user,
        )
        request = self.factory.post("/api/requirement-analysis/testcase-generation/generate/", {
            "title": "登录 PRD",
            "requirement_text": "手机号登录",
            "requirement_ids": ["REQ-1"],
            "case_type": "功能测试",
            "case_creator": "张三",
            "iteration": "2026.06",
            "use_writer_model": True,
            "use_reviewer_model": False,
        }, format="json")
        request.user = self.user

        response = TestCaseGenerationTaskViewSet.as_view({"post": "generate"})(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn("API Key", response.data["error"])


class PRD2CaseBackendAPITests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("api_refactor_user", password="password123")
        self.factory = APIRequestFactory()
        self.writer_config = AIModelConfig.objects.create(
            name="Writer",
            model_type="deepseek",
            role="writer",
            api_key="test-key",
            base_url="https://api.example.com/v1",
            model_name="writer-model",
            created_by=self.user,
        )
        self.prompt = PromptConfig.objects.create(
            name="Writer prompt",
            prompt_type="writer",
            content="你是测试专家",
            is_active=True,
            created_by=self.user,
        )

    def test_generate_accepts_uploaded_txt_and_template(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["用例名称", "用例状态", "创建人"])
        buffer = BytesIO()
        workbook.save(buffer)
        request = self.factory.post("/api/requirement-analysis/testcase-generation/generate/", {
            "title": "登录 PRD",
            "source_file": SimpleUploadedFile("prd.txt", b"login prd", content_type="text/plain"),
            "template_file": SimpleUploadedFile(
                "template.xlsx",
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "requirement_ids": "REQ-1",
            "case_type": "功能测试",
            "case_creator": "张三",
            "iteration": "2026.06",
        }, format="multipart")
        request.user = self.user

        with mock.patch.object(TestCaseGenerationTaskViewSet, "_start_test_point_generation_thread"):
            response = TestCaseGenerationTaskViewSet.as_view({"post": "generate"})(request)

        self.assertEqual(response.status_code, 201)
        task = TestCaseGenerationTask.objects.get(task_id=response.data["task_id"])
        self.assertEqual(task.requirement_text, "login prd")
        self.assertEqual(task.source_file_type, "txt")
        self.assertEqual(task.source_extract_status, "parsed")
        self.assertEqual(task.template_schema["headers"], ["用例名称", "用例状态", "创建人"])

    def test_generate_reuses_writer_model_for_image_when_vision_config_missing(self):
        image = Image.new("RGB", (4, 4), "white")
        image_buffer = BytesIO()
        image.save(image_buffer, format="PNG")
        request = self.factory.post("/api/requirement-analysis/testcase-generation/generate/", {
            "title": "图片 PRD",
            "source_file": SimpleUploadedFile("prd.png", image_buffer.getvalue(), content_type="image/png"),
            "requirement_ids": "REQ-IMG",
            "case_type": "功能测试",
            "case_creator": "张三",
            "iteration": "2026.06",
        }, format="multipart")
        request.user = self.user

        with mock.patch.object(TestCaseGenerationTaskViewSet, "_start_test_point_generation_thread"), \
             mock.patch(
                 "apps.requirement_analysis.views.VisionDocumentExtractor.extract_text",
                 new=AsyncMock(return_value="图片中的 PRD 文本"),
             ):
            response = TestCaseGenerationTaskViewSet.as_view({"post": "generate"})(request)

        self.assertEqual(response.status_code, 201)
        task = TestCaseGenerationTask.objects.get(task_id=response.data["task_id"])
        self.assertEqual(task.source_file_type, "png")
        self.assertEqual(task.requirement_text, "图片中的 PRD 文本")

    def test_generate_xmind_skips_test_point_generation_and_starts_case_generation(self):
        request = self.factory.post("/api/requirement-analysis/testcase-generation/generate/", {
            "title": "XMind 测试点",
            "source_file": SimpleUploadedFile(
                "login.xmind",
                make_xmind_bytes(),
                content_type="application/octet-stream",
            ),
            "requirement_ids": "REQ-XMIND",
            "case_type": "功能测试",
            "case_creator": "张三",
            "iteration": "2026.06",
        }, format="multipart")
        request.user = self.user

        with mock.patch.object(TestCaseGenerationTaskViewSet, "_start_test_point_generation_thread") as start_points, \
             mock.patch.object(TestCaseGenerationTaskViewSet, "_start_case_generation_thread") as start_cases:
            response = TestCaseGenerationTaskViewSet.as_view({"post": "generate"})(request)

        self.assertEqual(response.status_code, 201)
        task = TestCaseGenerationTask.objects.get(task_id=response.data["task_id"])
        self.assertEqual(task.source_file_type, "xmind")
        self.assertEqual(task.test_points_review_status, "approved")
        self.assertEqual(task.pipeline_artifacts["current_stage"], "case_generation")
        self.assertEqual(task.test_points[0]["requirement_ids"], ["REQ-XMIND"])
        self.assertEqual(task.test_points[0]["review_status"], "approved")
        start_points.assert_not_called()
        start_cases.assert_called_once_with(task.task_id)

    def test_revise_test_points_requires_message(self):
        task = TestCaseGenerationTask.objects.create(
            task_id="TASK_REV001",
            title="登录 PRD",
            requirement_text="手机号登录",
            requirement_ids=["REQ-1"],
            case_type="功能测试",
            case_creator="张三",
            iteration="2026.06",
            test_points=[{"id": "TP-1", "title": "旧点"}],
            writer_model_config=self.writer_config,
            writer_prompt_config=self.prompt,
            created_by=self.user,
        )
        request = self.factory.post(
            f"/api/requirement-analysis/testcase-generation/{task.task_id}/revise_test_points/",
            {},
            format="json",
        )
        request.user = self.user

        response = TestCaseGenerationTaskViewSet.as_view({"post": "revise_test_points"})(request, task_id=task.task_id)

        self.assertEqual(response.status_code, 400)
        self.assertIn("message", response.data["error"])

    def test_export_excel_returns_xlsx_after_approval(self):
        task = TestCaseGenerationTask.objects.create(
            task_id="TASK_XLSX001",
            title="登录 PRD",
            requirement_text="手机号登录",
            requirement_ids=["REQ-1"],
            case_type="功能测试",
            case_creator="张三",
            iteration="2026.06",
            test_cases_json=[{"title": "登录成功", "case_status": "已评审"}],
            test_cases_review_status="approved",
            template_schema={"headers": ["用例名称", "用例状态", "创建人"]},
            created_by=self.user,
        )
        request = self.factory.get(f"/api/requirement-analysis/testcase-generation/{task.task_id}/export_excel/")
        request.user = self.user

        response = TestCaseGenerationTaskViewSet.as_view({"get": "export_excel"})(request, task_id=task.task_id)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual([cell.value for cell in workbook.active[2]], ["登录成功", None, "张三"])


class PRD2CaseReviewGateTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("review_user", password="password123")
        self.project = Project.objects.create(
            name="PRD2Case 项目",
            owner=self.user,
        )
        self.task = TestCaseGenerationTask.objects.create(
            task_id="TASK_GATE001",
            title="登录 PRD",
            requirement_text="手机号登录",
            requirement_ids=["REQ-1"],
            case_type="功能测试",
            case_creator="张三",
            iteration="2026.06",
            test_points=[{"id": "TP-001", "title": "登录主流程", "review_status": "pending"}],
            project=self.project,
            created_by=self.user,
        )

    def test_export_requires_case_review_approval(self):
        factory = APIRequestFactory()
        request = factory.get(f"/api/requirement-analysis/testcase-generation/{self.task.task_id}/export_excel/")
        request.user = self.user

        response = TestCaseGenerationTaskViewSet.as_view({"get": "export_excel"})(request, task_id=self.task.task_id)

        self.assertEqual(response.status_code, 400)
        self.assertIn("审核", response.data["error"])

    def test_save_to_records_uses_approved_structured_cases(self):
        self.task.status = "completed"
        self.task.test_cases_review_status = "approved"
        self.task.test_cases_json = [{
            "title": "手机号验证码登录成功",
            "preconditions": ["用户未登录", "验证码有效"],
            "steps": [
                {"index": 1, "action": "输入手机号", "expected": "展示验证码输入框"},
                {"index": 2, "action": "输入验证码并提交", "expected": "登录成功"},
            ],
            "expected_result": "登录成功并进入首页",
            "priority": "P1",
        }]
        self.task.save()
        factory = APIRequestFactory()
        request = factory.post(
            f"/api/requirement-analysis/testcase-generation/{self.task.task_id}/save_to_records/",
            {},
            format="json",
        )
        request.user = self.user

        response = TestCaseGenerationTaskViewSet.as_view({"post": "save_to_records"})(request, task_id=self.task.task_id)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["imported_count"], 1)
        managed_case = ManagedTestCase.objects.get(project=self.project)
        self.assertEqual(managed_case.title, "手机号验证码登录成功")
        self.assertIn("输入手机号", managed_case.steps)
        self.assertEqual(managed_case.priority, "high")
