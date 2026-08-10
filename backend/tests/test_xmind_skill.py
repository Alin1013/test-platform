import asyncio
import io
import json
import zipfile

import httpx
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.services.xmind_skill import STANDARD_HEADERS, align_generated_cases
from backend.app.services.xmind_skill import (
    LLMConfig,
    MAX_GENERATED_CASES,
    XMindGenerationError,
    XMindToTestCaseSkill,
)


def make_grouped_xmind_file() -> bytes:
    content = [
        {
            "title": "用户中心",
            "rootTopic": {
                "title": "用户中心",
                "children": {
                    "attached": [
                        {
                            "title": "注册模块",
                            "children": {
                                "attached": [{"title": "注册成功"}],
                            },
                        },
                        {
                            "title": "登录模块",
                            "children": {
                                "attached": [{"title": "密码错误"}],
                            },
                        },
                    ]
                },
            },
        }
    ]
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("content.json", json.dumps(content, ensure_ascii=False))
    return buffer.getvalue()


def llm_response(cases: list[dict]) -> httpx.Response:
    return httpx.Response(
        200,
        json={
            "choices": [
                {"message": {"content": json.dumps({"cases": cases}, ensure_ascii=False)}}
            ]
        },
    )


def test_generated_cases_are_aligned_to_all_required_headers() -> None:
    cases = align_generated_cases(
        [{"用例名称": "注册成功", "用例等级": "高"}],
        directory="用户中心/注册模块",
        creator="江珊",
    )

    assert list(cases[0]) == list(STANDARD_HEADERS)
    assert cases[0]["用例目录"] == "用户中心/注册模块"
    assert cases[0]["用例状态"] == "草稿"
    assert cases[0]["用例等级"] == "P0"
    assert cases[0]["创建人"] == "江珊"
    assert cases[0]["前置条件"] == "无"
    assert cases[0]["用例步骤"].startswith("1.")
    assert cases[0]["预期结果"].startswith("1.")


def test_generated_case_directory_is_fixed_to_the_xmind_group() -> None:
    cases = align_generated_cases(
        [{"用例目录": "其他模块/错误目录", "用例名称": "注册成功"}],
        directory="用户中心/注册模块",
        creator="江珊",
    )

    assert cases[0]["用例目录"] == "用户中心/注册模块"


def test_generated_preview_can_be_confirmed_and_exported(client: TestClient) -> None:
    cases = [
        {
            "用例目录": "用户中心/注册模块",
            "用例名称": "注册成功",
            "需求ID": "REQ-1",
            "前置条件": "用户未注册",
            "用例类型": "功能测试",
            "用例状态": "草稿",
            "用例等级": "P1",
            "创建人": "江珊",
            "归属迭代": "Sprint 1",
            "用例步骤": "1. 提交注册信息",
            "预期结果": "1. 注册成功",
        }
    ]
    confirmed = client.post(
        "/api/v1/xmind/confirm",
        json={"uploader_id": 1, "module_mapping": {"用户中心/注册模块": "auth"}, "cases": cases},
    )

    assert confirmed.status_code == 201
    assert confirmed.json()["saved_cases"][0]["title"] == "注册成功"
    assert client.get("/api/v1/test-cases", params={"keyword": "注册成功"}).json()["total"] == 1

    exported = client.post("/api/v1/xmind/export", json={"cases": cases})
    assert exported.status_code == 200
    workbook = load_workbook(io.BytesIO(exported.content), read_only=True)
    assert list(next(workbook.active.iter_rows(values_only=True))) == list(STANDARD_HEADERS)


def test_xmind_skill_retries_unexpected_client_errors_and_caps_attempts() -> None:
    calls = 0

    class FailingClient:
        async def complete(self, **_: str) -> str:
            nonlocal calls
            calls += 1
            raise RuntimeError("connection reset")

    skill = XMindToTestCaseSkill(
        FailingClient(),
        max_attempts=20,
        retry_delay_seconds=0,
    )

    with pytest.raises(XMindGenerationError, match="XMind 用例生成失败，请稍后重试"):
        asyncio.run(
            skill.generate(
                [{"title": "登录", "children": []}],
                config=LLMConfig(api_key="key", base_url="https://llm.example", model="model"),
                creator="测试员",
            )
        )

    assert calls == 3


def test_xmind_skill_bounds_group_concurrency() -> None:
    active = 0
    peak = 0

    class TrackingClient:
        async def complete(self, **_: str) -> str:
            nonlocal active, peak
            active += 1
            peak = max(peak, active)
            try:
                await asyncio.sleep(0.01)
                return json.dumps({"cases": [{"用例名称": "有效用例"}]}, ensure_ascii=False)
            finally:
                active -= 1

    skill = XMindToTestCaseSkill(
        TrackingClient(),
        max_concurrency=2,
        retry_delay_seconds=0,
    )
    cases = asyncio.run(
        skill.generate(
            [
                {
                    "title": "用户中心",
                    "children": [
                        {"title": f"功能 {index}", "children": []}
                        for index in range(6)
                    ],
                }
            ],
            config=LLMConfig(api_key="key", base_url="https://llm.example", model="model"),
            creator="测试员",
        )
    )

    assert len(cases) == 6
    assert peak == 2


def test_xmind_skill_rejects_preview_larger_than_confirmation_limit() -> None:
    class OversizedClient:
        async def complete(self, **_: str) -> str:
            return json.dumps(
                {"cases": [{"用例名称": "有效用例"}] * (MAX_GENERATED_CASES + 1)},
                ensure_ascii=False,
            )

    skill = XMindToTestCaseSkill(OversizedClient(), retry_delay_seconds=0)

    with pytest.raises(XMindGenerationError, match="超过 5000 条限制"):
        asyncio.run(
            skill.generate(
                [{"title": "登录", "children": []}],
                config=LLMConfig(api_key="key", base_url="https://llm.example", model="model"),
                creator="测试员",
            )
        )
