"""处理用例文件导入导出，并保证批量导入的事务一致性。"""

import csv
import io
import json
from dataclasses import dataclass
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from ..case_file_schemas import TestCaseExportRequest
from ..domain_defaults import DEFAULT_PROJECT_NAME
from ..models import Module, TestCase, User
from ..schemas import ApiDetailsCreate, TestCaseCreate, UiDetailsCreate
from . import test_cases
from .xmind_skill import STANDARD_HEADERS

MAX_IMPORT_BYTES = 10 * 1024 * 1024
EXPORT_HEADERS = (
    "code",
    "title",
    "type",
    "module_id",
    "priority",
    "status",
    "author_id",
    "url",
    "method",
    "expected_code",
    "headers",
    "request_body",
    "expected_response",
    "steps",
    "requirement_id",
    "precondition",
    "test_steps",
    "expected_result",
    "iteration",
    "is_smoke",
    "project_name",
)
FUNCTIONAL_HEADERS = (
    "用例目录", "用例名称", "需求ID", "前置条件", "用例步骤", "预期结果",
    "用例类型", "用例状态", "用例等级", "创建人", "归属迭代", "是否冒烟", "项目归属",
)
HEADER_ALIASES = {
    "编号": "code",
    "用例名称": "title",
    "类型": "type",
    "模块ID": "module_id",
    "模块": "module_id",
    "模块名称": "module_id",
    "所属模块": "module_id",
    "所属模块名称": "module_id",
    "优先级": "priority",
    "状态": "status",
    "创建人ID": "author_id",
    "接口地址": "url",
    "HTTP方法": "method",
    "预期状态码": "expected_code",
    "请求头": "headers",
    "请求体": "request_body",
    "预期响应": "expected_response",
    "测试步骤": "steps",
    "用例目录": "module_id",
    "需求ID": "requirement_id",
    "前置条件": "precondition",
    "用例步骤": "test_steps",
    "预期结果": "expected_result",
    "用例类型": "type",
    "用例状态": "status",
    "用例等级": "priority",
    "创建人": "author_id",
    "归属迭代": "iteration",
    "是否冒烟": "is_smoke",
    "项目归属": "project_name",
}
TYPE_ALIASES = {
    "功能用例": "functional",
    "功能测试": "functional",
    "功能测试用例": "functional",
    "接口用例": "api",
    "接口测试": "api",
    "接口测试用例": "api",
    "UI自动化": "ui",
    "UI测试": "ui",
    "UI测试用例": "ui",
}
PRIORITY_ALIASES = {
    "高": "P0",
    "最高": "P0",
    "P0": "P0",
    "中": "P1",
    "P1": "P1",
    "低": "P2",
    "P2": "P2",
    "很低": "P3",
    "极低": "P3",
    "最低": "P3",
    "P3": "P3",
}
STATUS_ALIASES = {
    "正常": "维护中",
    "启用": "维护中",
    "维护中": "维护中",
    "已通过": "已通过",
    "草稿": "草稿",
    "失败": "已失败",
    "已失败": "已失败",
    "停用": "已停用",
    "禁用": "已停用",
    "已停用": "已停用",
}


@dataclass
class ExportedFile:
    content: bytes
    media_type: str
    filename: str


def _json_cell(value: Any) -> str:
    """把 JSON 字段序列化为单元格文本，空值输出为空串。"""
    return "" if value is None else json.dumps(value, ensure_ascii=False)


def _export_row(test_case: dict) -> list[Any]:
    """把单个用例序列化为导出行（API/UI 详情合并为 JSON 列）。"""
    api = test_case["api_details"] or {}
    ui = test_case["ui_details"] or {}
    return [
        test_case["code"],
        test_case["title"],
        test_case["type"],
        test_case["module_id"],
        test_case["priority"],
        test_case["status"],
        test_case["author_id"],
        api.get("url", ""),
        api.get("method", ""),
        api.get("expected_code", ""),
        _json_cell(api.get("headers")),
        _json_cell(api.get("request_body")),
        _json_cell(api.get("expected_response")),
        _json_cell(ui.get("steps")),
        test_case.get("requirement_id", ""),
        test_case.get("precondition", ""),
        test_case.get("test_steps", ""),
        test_case.get("expected_result", ""),
        test_case.get("iteration", ""),
        "是" if test_case.get("is_smoke") else "否",
        test_case.get("project_name", DEFAULT_PROJECT_NAME),
    ]


def export_cases(
    session: Session, payload: TestCaseExportRequest
) -> ExportedFile:
    """按筛选条件导出用例为 CSV 或 XLSX 文件。"""
    query = test_cases.filtered_case_query(
        case_type=payload.type,
        module_id=payload.module_id,
        priority=payload.priority,
        status=payload.status,
        keyword=payload.keyword,
    )
    rows = session.scalars(query.order_by(TestCase.id)).all()
    exported_rows = [_export_row(test_cases.serialize_case(row)) for row in rows]

    if payload.format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(EXPORT_HEADERS)
        writer.writerows(exported_rows)
        return ExportedFile(
            # 添加 UTF-8 BOM，避免常见桌面版 Excel 将中文按本地编码解析。
            content=("\ufeff" + output.getvalue()).encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            filename="test-cases.csv",
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Cases"
    sheet.append(EXPORT_HEADERS)
    for row in exported_rows:
        sheet.append(row)
    output_bytes = io.BytesIO()
    workbook.save(output_bytes)
    return ExportedFile(
        content=output_bytes.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="test-cases.xlsx",
    )


def export_generated_cases(cases: list[dict[str, Any]]) -> ExportedFile:
    """导出尚未入库的 XMind 生成预览，表头顺序与 Skill 契约一致。"""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "XMind Generated Cases"
    sheet.append(STANDARD_HEADERS)
    for case in cases:
        sheet.append([case.get(header, "") for header in STANDARD_HEADERS])
    output_bytes = io.BytesIO()
    workbook.save(output_bytes)
    return ExportedFile(
        content=output_bytes.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="xmind-generated-cases.xlsx",
    )


def _csv_rows(content: bytes) -> list[dict[str, Any]]:
    """解析 CSV 字节流为字典行；要求 UTF-8 编码。"""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise HTTPException(status_code=422, detail="CSV must be UTF-8 encoded") from error
    return list(csv.DictReader(io.StringIO(text)))


def _xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    """解析 XLSX 工作簿：首行为表头，其余为数据行。"""
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise HTTPException(status_code=422, detail="Invalid XLSX workbook") from error
    values = workbook.active.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(values)]
    except StopIteration:
        return []
    return [dict(zip(headers, row, strict=False)) for row in values]


def _xls_rows(content: bytes) -> list[dict[str, Any]]:
    """解析旧版 XLS 工作簿（依赖 xlrd），首行为表头。"""
    try:
        import xlrd
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
    except Exception as error:
        raise HTTPException(status_code=422, detail="Invalid XLS workbook") from error
    if sheet.nrows == 0:
        return []
    headers = [str(value or "").strip() for value in sheet.row_values(0)]
    return [dict(zip(headers, sheet.row_values(index), strict=False)) for index in range(1, sheet.nrows)]


def _parse_json(value: Any, field: str) -> Any:
    """把单元格中的 JSON 字符串解析为对象，非法时报字段错误。"""
    if value is None or value == "":
        return None
    if not isinstance(value, str):
        return value
    try:
        return json.loads(value)
    except json.JSONDecodeError as error:
        raise ValueError(f"{field} must contain valid JSON") from error


def _normalize_alias(value: Any, aliases: dict[str, str], default: str | None = None) -> str:
    """把导入表中的别名（如“高”）归一化为内部枚举值。"""
    text = str(value or "").strip()
    if not text:
        return default or ""
    return aliases.get(text, text)


def _canonical_row(raw_row: dict[str, Any]) -> dict[str, Any]:
    """把中文表头映射为内部字段名。"""
    return {
        HEADER_ALIASES.get(str(key).strip(), str(key).strip()): value
        for key, value in raw_row.items()
    }


def _case_payload(raw_row: dict[str, Any]) -> TestCaseCreate:
    """从导入行构造用例创建载荷；API/UI 类型按列拼接详情。"""
    row = _canonical_row(raw_row)
    case_type = _normalize_alias(row.get("type"), TYPE_ALIASES)
    common = {
        "code": row.get("code") or None,
        "title": str(row.get("title") or "").strip(),
        "type": case_type,
        "module_id": str(row.get("module_id") or "").strip(),
        "priority": _normalize_alias(row.get("priority"), PRIORITY_ALIASES, "P1"),
        "status": _normalize_alias(row.get("status"), STATUS_ALIASES, "草稿"),
        "author_id": int(row.get("author_id") or 1),
        "requirement_id": str(row.get("requirement_id") or "").strip() or None,
        "precondition": str(row.get("precondition") or "").strip(),
        "test_steps": str(row.get("test_steps") or row.get("steps") or "").strip(),
        "expected_result": str(row.get("expected_result") or "").strip(),
        "iteration": str(row.get("iteration") or "").strip(),
        "is_smoke": str(row.get("is_smoke") or "").strip().lower() in {"是", "yes", "true", "1", "y"},
        "project_name": str(row.get("project_name") or DEFAULT_PROJECT_NAME).strip(),
    }
    if case_type == "api":
        common["api_details"] = ApiDetailsCreate(
            url=str(row.get("url") or "").strip(),
            method=str(row.get("method") or "GET").strip().upper(),
            expected_code=int(row.get("expected_code") or 200),
            headers=_parse_json(row.get("headers"), "headers") or {},
            request_body=_parse_json(row.get("request_body"), "request_body"),
            expected_response=_parse_json(row.get("expected_response"), "expected_response"),
        )
    elif case_type == "ui":
        common["ui_details"] = UiDetailsCreate(
            steps=_parse_json(row.get("steps"), "steps") or []
        )
    return TestCaseCreate.model_validate(common)


def import_cases(session: Session, filename: str, content: bytes, module_id: str | None = None) -> dict:
    """批量导入用例：整批单事务，任一行失败即整体回滚并定位行号。"""
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=413, detail="Import file exceeds the 10 MB limit")
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        rows = _csv_rows(content)
    elif lower_name.endswith(".xlsx"):
        rows = _xlsx_rows(content)
    elif lower_name.endswith(".xls"):
        rows = _xls_rows(content)
    else:
        raise HTTPException(status_code=415, detail="Only .csv, .xls and .xlsx files are supported")
    if not rows:
        raise HTTPException(status_code=422, detail="Import file has no data rows")

    raw_headers = {str(key).strip() for key in rows[0]}
    if any(header in raw_headers for header in FUNCTIONAL_HEADERS):
        if tuple(str(key).strip() for key in rows[0]) != FUNCTIONAL_HEADERS:
            raise HTTPException(status_code=422, detail="功能用例导入表头不一致，请使用标准模板")

    # 所有行共用一个事务；任意一行失败时，前面已 flush 的记录也会回滚。
    created = []
    selected_module_id = module_id.strip() if module_id and module_id.strip() else None
    for index, row in enumerate(rows, start=2):
        try:
            row = _canonical_row(row)
            if selected_module_id:
                row = {**row, "module_id": selected_module_id}
            normalized_module = str(row.get("module_id") or "").strip()
            if not normalized_module:
                raise ValueError(
                    "module_id is required; fill in the module column or select a module"
                )
            if normalized_module:
                module = session.get(Module, normalized_module)
                if module is None:
                    module = session.scalar(select(Module).where(Module.name == normalized_module))
                if module is not None:
                    row = {**row, "module_id": module.id}
            author_value = row.get("创建人") or row.get("author_id")
            if author_value and not str(author_value).strip().isdigit():
                author = session.scalar(select(User).where(User.name == str(author_value).strip()))
                if author is None:
                    raise ValueError(f"创建人不存在: {author_value}")
                row = {**row, "author_id": author.id}
            payload = _case_payload(row)
            created.append(test_cases.add_case(session, payload))
        except (ValueError, ValidationError, HTTPException, IntegrityError) as error:
            session.rollback()
            detail = error.detail if isinstance(error, HTTPException) else str(error)
            raise HTTPException(status_code=422, detail=f"Invalid row {index}: {detail}") from error
    try:
        session.commit()
    except IntegrityError as error:
        session.rollback()
        raise HTTPException(status_code=409, detail="Imported case code already exists") from error
    return {
        "imported_count": len(created),
        "codes": [test_case.code for test_case in created],
    }
